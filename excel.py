"""
Excel (.xlsx) dışa aktarma katmanı.

Rapor sayfalarındaki tablolar buradan Excel dosyasına çevrilir. Amaç, depo
sorumlusunun listeyi Excel'de açıp filtreleyebilmesi, yazdırabilmesi veya
muhasebe/satın alma birimine gönderebilmesi.

Tasarım notu: dosya diske yazılmaz, bellekte (BytesIO) oluşturulup doğrudan
tarayıcıya gönderilir. Böylece sunucuda geçici dosya birikmez.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Arayüzdeki lacivert/gri paletiyle uyumlu stiller
BASLIK_DOLGU = PatternFill("solid", fgColor="1E3A5F")
BASLIK_YAZI = Font(color="FFFFFF", bold=True, size=11)
ANA_BASLIK_YAZI = Font(color="1E3A5F", bold=True, size=14)
FILTRE_YAZI = Font(color="6B7280", size=10, italic=True)
INCE_KENAR = Border(bottom=Side(style="thin", color="E5E7EB"))

TARIH_BICIMI = "dd.mm.yyyy hh:mm"


def tabloyu_excele_cevir(baslik, sutunlar, satirlar, filtre_notu=""):
    """Bir tabloyu .xlsx dosyasına çevirip bayt olarak döner.

    baslik      : sayfanın üst başlığı (örn. "Mal Giriş Raporu")
    sutunlar    : [(başlık, alan_adı, genişlik, tip), ...]
                  tip: "metin" | "sayi" | "tarih"
    satirlar    : sqlite3.Row listesi
    filtre_notu : hangi filtrelerle alındığı (raporun altına yazılır)
    """
    kitap = Workbook()
    sayfa = kitap.active
    sayfa.title = baslik[:31]        # Excel sayfa adı en fazla 31 karakter

    sutun_sayisi = len(sutunlar)

    # --- Üst bilgi: başlık, alınma zamanı, uygulanan filtreler ---
    sayfa.cell(row=1, column=1, value=baslik).font = ANA_BASLIK_YAZI
    sayfa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sutun_sayisi)

    alinma = datetime.now().strftime("%d.%m.%Y %H:%M")
    ust_not = f"Rapor tarihi: {alinma}"
    if filtre_notu:
        ust_not += f"   |   Filtre: {filtre_notu}"
    sayfa.cell(row=2, column=1, value=ust_not).font = FILTRE_YAZI
    sayfa.merge_cells(start_row=2, start_column=1, end_row=2, end_column=sutun_sayisi)

    BASLIK_SATIRI = 4

    # --- Sütun başlıkları ---
    for i, (sutun_basligi, _, genislik, _) in enumerate(sutunlar, start=1):
        hucre = sayfa.cell(row=BASLIK_SATIRI, column=i, value=sutun_basligi)
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(vertical="center")
        sayfa.column_dimensions[get_column_letter(i)].width = genislik
    sayfa.row_dimensions[BASLIK_SATIRI].height = 22

    # --- Veri satırları ---
    for satir_no, satir in enumerate(satirlar, start=BASLIK_SATIRI + 1):
        for i, (_, alan, _, tip) in enumerate(sutunlar, start=1):
            deger = satir[alan] if alan in satir.keys() else None
            hucre = sayfa.cell(row=satir_no, column=i, value=_donustur(deger, tip))
            hucre.border = INCE_KENAR
            if tip == "tarih":
                hucre.number_format = TARIH_BICIMI
            elif tip == "sayi":
                hucre.number_format = "#,##0.##"

    # --- Excel'in kendi filtre/dondurma özellikleri ---
    son_satir = BASLIK_SATIRI + len(satirlar)
    if satirlar:
        sayfa.auto_filter.ref = (f"A{BASLIK_SATIRI}:"
                                 f"{get_column_letter(sutun_sayisi)}{son_satir}")
    # Başlık satırı kaydırınca sabit kalsın
    sayfa.freeze_panes = f"A{BASLIK_SATIRI + 1}"

    # --- Alt bilgi ---
    sayfa.cell(row=son_satir + 2, column=1,
               value=f"Toplam {len(satirlar)} kayıt").font = FILTRE_YAZI

    tampon = BytesIO()
    kitap.save(tampon)
    tampon.seek(0)
    return tampon


def _donustur(deger, tip):
    """Veritabanından gelen değeri Excel hücresine uygun tipe çevirir.
    Tarihler metin olarak saklandığı için gerçek datetime'a çevrilir; böylece
    Excel'de tarih olarak sıralanıp filtrelenebilir."""
    if deger is None or deger == "":
        return "" if tip != "sayi" else 0
    if tip == "tarih":
        try:
            return datetime.strptime(str(deger)[:19], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return str(deger)
    if tip == "sayi":
        try:
            return float(deger)
        except (TypeError, ValueError):
            return 0
    return str(deger)


def dosya_adi(on_ek):
    """Zaman damgalı dosya adı — indirilen dosyalar birbirini ezmesin."""
    return f"{on_ek}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


# ============================================================
# RAPOR SÜTUN TANIMLARI
# (başlık, veri alanı, sütun genişliği, tip)
# ============================================================

GIRIS_SUTUNLARI = [
    ("Ne zaman",      "tarih",       18, "tarih"),
    ("Stok kodu",     "stok_kodu",   16, "metin"),
    ("Malzeme",       "urun_ad",     32, "metin"),
    ("Miktar",        "miktar",      10, "sayi"),
    ("Birim",         "birim",        9, "metin"),
    ("Bölüm",         "bolum_ad",    28, "metin"),
    ("Raf",           "raf_kod",     12, "metin"),
    ("Kim koydu",     "calisan_ad",  20, "metin"),
    ("Tedarikçi",     "tedarikci",   24, "metin"),
    ("İrsaliye no",   "irsaliye_no", 16, "metin"),
    ("Açıklama",      "aciklama",    30, "metin"),
]

CIKIS_SUTUNLARI = [
    ("Ne zaman",         "tarih",       18, "tarih"),
    ("Stok kodu",        "stok_kodu",   16, "metin"),
    ("Malzeme",          "urun_ad",     32, "metin"),
    ("Miktar",           "miktar",      10, "sayi"),
    ("Birim",            "birim",        9, "metin"),
    ("Bölüm (nereden)",  "bolum_ad",    28, "metin"),
    ("Raf",              "raf_kod",     12, "metin"),
    ("Birim kodu",       "birim_kod",   12, "metin"),
    ("Hangi birim için", "birim_ad",    28, "metin"),
    ("Neden",            "neden",       22, "metin"),
    ("Kim çıkardı",      "calisan_ad",  20, "metin"),
    ("Teslim alan",      "teslim_alan", 20, "metin"),
    ("Açıklama",         "aciklama",    30, "metin"),
]

HAREKET_SUTUNLARI = [
    ("Tarih",          "tarih",         18, "tarih"),
    ("Tip",            "tip",            9, "metin"),
    ("Malzeme",        "urun_ad",       32, "metin"),
    ("Miktar",         "miktar",        10, "sayi"),
    ("Birim",          "birim",          9, "metin"),
    ("Bölüm",          "bolum_ad",      28, "metin"),
    ("Raf",            "raf_kod",       12, "metin"),
    ("Fabrika birimi", "birim_ad",      28, "metin"),
    ("Kaynak / Neden", "kaynak_hedef",  24, "metin"),
    ("Kaydeden",       "calisan_ad",    20, "metin"),
    ("Detay",          "detay",         30, "metin"),
]

STOK_SUTUNLARI = [
    ("Bölüm",     "bolum_ad",  28, "metin"),
    ("Raf",       "raf_kod",   12, "metin"),
    ("Stok kodu", "stok_kodu", 16, "metin"),
    ("Malzeme",   "urun_ad",   32, "metin"),
    ("Miktar",    "stok",      12, "sayi"),
    ("Birim",     "birim",      9, "metin"),
]
